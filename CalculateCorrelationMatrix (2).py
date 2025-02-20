from PySide6.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QLabel, 
                               QTableWidget, QTableWidgetItem, QSplitter,
                               QWidget, QLineEdit,   QAbstractItemView, QPushButton, QHeaderView, QComboBox, QListWidget, QMessageBox, QCheckBox, QFileDialog)
from PySide6.QtCore import Qt
from PySide6.QtGui import QColor, QDoubleValidator, QPalette
from StyledTwoListSelector import TwoListSelector
from StyledTwoListSelector import DarkTwoListSelector
from PySide6.QtWidgets import QStackedWidget, QMenu, QSlider
import networkx as nx
import matplotlib.colors as mcolors

import pandas as pd
import seaborn as sns
from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg
import matplotlib.colors as mcolors
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import PatternFill
import numpy as np
import traceback


class SaveRegressionDialog(QDialog):
    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.setWindowTitle("Save to Regression Table")
        self.setMinimumWidth(400)
        
        layout = QVBoxLayout()
        
        # Table selection/creation section
        layout.addWidget(QLabel("Select or Create Table:"))
        self.table_combo = QComboBox()
        self.table_combo.setEditable(True)
        self.table_combo.setInsertPolicy(QComboBox.InsertPolicy.InsertAtBottom)
        self.populate_regression_tables()
        layout.addWidget(self.table_combo)
        
        # Save mode selection
        layout.addWidget(QLabel("Save Mode:"))
        self.mode_combo = QComboBox()
        self.mode_combo.addItems(["Replace Regression", "Add to Regression"])
        layout.addWidget(self.mode_combo)
        
        # Description field
        layout.addWidget(QLabel("Description:"))
        self.description_input = QLineEdit()
        self.description_input.setPlaceholderText("Enter table description...")
        layout.addWidget(self.description_input)
        
        # When existing table is selected, show its description
        self.table_combo.currentTextChanged.connect(self.load_table_description)
        
        # Buttons
        button_layout = QHBoxLayout()
        save_button = QPushButton("Save")
        cancel_button = QPushButton("Cancel")
        
        save_button.clicked.connect(self.accept)
        cancel_button.clicked.connect(self.reject)
        
        button_layout.addWidget(save_button)
        button_layout.addWidget(cancel_button)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def populate_regression_tables(self):
        """Populate the combo box with existing regression entries"""
        tables = self.db_manager.get_regression_tables()
        self.table_combo.clear()
        self.table_combo.addItem("")  # Empty option for new tables
    
        for regression in tables:
            # Since get_regression_tables now returns more fields, unpack them
            regression_id, regression_name, values_table, attrs_table, description, date_created = regression
        
            # Add regression name to combo box
            # You could also store the full regression info in the item's data
            self.table_combo.addItem(regression_name)
    
    def load_table_description(self, table_name):
        """Load description when a table is selected"""
        if table_name:
            table_info = self.db_manager.get_regression_table_by_name(table_name)
            if table_info and table_info[1]:  # If description exists
                self.description_input.setText(table_info[1])
            else:
                self.description_input.clear()
        else:
            self.description_input.clear()
    
    def get_selected_table(self):
        """Return the selected or new table name"""
        return self.table_combo.currentText().strip()
    
    def get_description(self):
        """Return the description"""
        return self.description_input.text().strip()
    
    def get_save_mode(self):
        """Return whether to replace or add to regression"""
        return self.mode_combo.currentText() == "Replace Regression"

class CrossplotDialog(QDialog):
    def __init__(self, data_matrix, x_col, y_col, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Crossplot: {x_col} vs {y_col}")
        self.setMinimumSize(800, 600)
        
        # Set up a dark palette
        dark_palette = QPalette()
        dark_palette.setColor(QPalette.Window, QColor(0, 0, 0))
        dark_palette.setColor(QPalette.WindowText, QColor(255, 255, 255))
        dark_palette.setColor(QPalette.Base, QColor(0, 0, 0))
        dark_palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
        dark_palette.setColor(QPalette.ToolTipBase, QColor(255, 255, 255))
        dark_palette.setColor(QPalette.ToolTipText, QColor(255, 255, 255))
        dark_palette.setColor(QPalette.Text, QColor(255, 255, 255))
        dark_palette.setColor(QPalette.Button, QColor(53, 53, 53))
        dark_palette.setColor(QPalette.ButtonText, QColor(255, 255, 255))
        dark_palette.setColor(QPalette.BrightText, QColor(255, 0, 0))
        dark_palette.setColor(QPalette.Highlight, QColor(142, 45, 197).lighter())
        dark_palette.setColor(QPalette.HighlightedText, QColor(0, 0, 0))
        
        # Apply the palette to the dialog
        self.setPalette(dark_palette)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        
        fig, ax = plt.subplots(figsize=(8, 6))
        canvas = FigureCanvasQTAgg(fig)
        
        # Set figure and axes background to black
        fig.patch.set_facecolor('black')
        ax.set_facecolor('black')
        
        sns.regplot(
            x=self.data_matrix[x_col], 
            y=self.data_matrix[y_col], 
            scatter_kws={'alpha': 0.7, 'color': '#00BFFF', 's': 30},  # Bright blue scatter points
            line_kws={'color': '#e74c3c', 'linewidth': 1},  # Keep the red regression line
            ax=self.crossplot_ax
        )
        
        ax.set_xlabel(x_col, color='white')
        ax.set_ylabel(y_col, color='white')
        ax.set_title(f"Crossplot: {x_col} vs {y_col}", color='white')
        ax.spines['bottom'].set_color('white')
        ax.spines['top'].set_color('white')
        ax.spines['right'].set_color('white')
        ax.spines['left'].set_color('white')
        ax.tick_params(axis='x', colors='white')
        ax.tick_params(axis='y', colors='white')
        
        corr = data_matrix[x_col].corr(data_matrix[y_col])
        ax.annotate(f'Correlation: {corr:.2f}', 
                    xy=(0.05, 0.95), 
                    xycoords='axes fraction',
                    fontsize=10,
                    color='white',
                    verticalalignment='top')
        
        fig.tight_layout()
        layout.addWidget(canvas)
        
        # Ensure the dialog uses the dark background
        self.setAutoFillBackground(True)







class CorrelationDisplayDialog(QDialog):
    def __init__(self, correlation_matrix, data_matrix, selected_attrs, threshold, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Correlation Matrix Results")
        self.setMinimumWidth(1400)
        self.setMinimumHeight(900)

        self.correlation_matrix = correlation_matrix  # Already filtered by threshold!
        self.data_matrix = data_matrix
        self.selected_attrs = selected_attrs
        self.threshold = threshold  # Store threshold

        self.setWindowFlags(self.windowFlags() | Qt.WindowMaximizeButtonHint)

        # 🔹 Apply dark theme to remove white borders
        self.setStyleSheet("""
            QDialog {
                background-color: #2c2c2c;
                border: none;
            }
            QWidget {
                background-color: #2c2c2c;
                color: #e0e0e0;
            }
            QLabel {
                color: #e0e0e0;
            }
            QSplitter::handle {
                background-color: #444;
            }
            QTableWidget {
                background-color: #3c3f41;
                color: #e0e0e0;
                border: 1px solid #5a5a5a;
            }
            QHeaderView::section {
                background-color: #2c2c2c;
                color: #e0e0e0;
                padding: 4px;
                border: 1px solid #5a5a5a;
                font-weight: bold;
            }
            QPushButton {
                background-color: #4a6984;
                color: white;
                border: 1px solid #5a5a5a;
                border-radius: 4px;
                padding: 6px 12px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #5a7a9d;
            }
            QPushButton:pressed {
                background-color: #3a5a7d;
            }
            QLineEdit {
                background-color: #3c3f41;
                color: #e0e0e0;
                border: 1px solid #5a5a5a;
                border-radius: 4px;
                padding: 4px;
            }
            QLineEdit:focus {
                border: 1px solid #4a6984;
            }
            QComboBox {
                background-color: #3c3f41;
                color: #e0e0e0;
                border: 1px solid #5a5a5a;
                border-radius: 4px;
                padding: 2px;
            }
            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 15px;
                border-left-width: 1px;
                border-left-color: #5a5a5a;
                border-left-style: solid;
            }
            QListWidget {
                background-color: #3c3f41;
                color: #e0e0e0;
                border: 1px solid #5a5a5a;
                border-radius: 4px;
                padding: 2px;
            }
            QListWidget::item {
                background-color: #3c3f41;
                padding: 2px;
                border-bottom: 1px solid #4a4a4a;
            }
            QListWidget::item:selected {
                background-color: #4a6984;
                color: white;
            }
            QListWidget::item:hover {
                background-color: #4a4a4a;
            }
        """)

        # Main Layout: Horizontal Splitter
        main_layout = QHBoxLayout(self)
        splitter = QSplitter(Qt.Horizontal)
        
        # Left Panel: Heatmap & Crossplot (Stacked)
        left_panel = self.create_left_panel()
        splitter.addWidget(left_panel)

        # Right Panel: Correlation Threshold, Attribute Selection, and Data Table
        right_panel = self.create_right_panel()
        splitter.addWidget(right_panel)

        self.populate_attributes()
        # Set stretch factors for better spacing
        splitter.setStretchFactor(0, 3)  # Heatmap & Crossplot
        splitter.setStretchFactor(1, 2)  # Right Panel

        main_layout.addWidget(splitter)

        # Draw Initial Heatmap
        self.draw_heatmap()
        self.draw_network()

    def create_left_panel(self):
        """Create left panel with visualization controls"""
        left_panel = QWidget()
        left_panel.setStyleSheet("""
            QWidget {
                background-color: #2c2c2c;
                color: #e0e0e0;
                font-size: 10px;
            }
            QLabel {
                color: #e0e0e0;
                font-size: 10px;
            }
            QComboBox {
                background-color: #3c3f41;
                color: #e0e0e0;
                border: 1px solid #5a5a5a;
                border-radius: 4px;
                padding: 2px;
                font-size: 10px;
            }
            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 15px;
                border-left-width: 1px;
                border-left-color: #5a5a5a;
                border-left-style: solid;
            }
        """)
    
        layout = QVBoxLayout(left_panel)
    
        # View selector
        view_layout = QHBoxLayout()
        view_label = QLabel("View:")
        view_label.setStyleSheet("font-size: 10px;")
        self.view_selector = QComboBox()
        self.view_selector.addItems(["Correlation Matrix", "Network Graph"])
        self.view_selector.currentTextChanged.connect(self.switch_view)
        view_layout.addWidget(view_label)
        view_layout.addWidget(self.view_selector)
        layout.addLayout(view_layout)
    
        # Create stacked widget for visualizations
        self.viz_stack = QStackedWidget()

        # Heatmap widget
        self.heatmap_widget = QWidget()
        heatmap_layout = QVBoxLayout(self.heatmap_widget)
        self.heatmap_fig = Figure(figsize=(8, 6), facecolor='#2c2c2c', edgecolor='none', dpi=100)
        self.heatmap_canvas = FigureCanvasQTAgg(self.heatmap_fig)
        self.heatmap_ax = self.heatmap_fig.add_subplot(111)
        heatmap_layout.addWidget(self.heatmap_canvas)

        # Network widget
        self.network_widget = QWidget()
        network_layout = QVBoxLayout(self.network_widget)
        self.network_fig = Figure(figsize=(8, 6), facecolor='#2c2c2c', edgecolor='none', dpi=100)
        self.network_canvas = FigureCanvasQTAgg(self.network_fig)
        self.network_ax = self.network_fig.add_subplot(111)
        network_layout.addWidget(self.network_canvas)

        # Add both widgets to stack
        self.viz_stack.addWidget(self.heatmap_widget)
        self.viz_stack.addWidget(self.network_widget)
        layout.addWidget(self.viz_stack)
    
        self.crossplot_fig = Figure(figsize=(6, 5), facecolor='#2c2c2c', edgecolor='none', dpi=100)
        self.crossplot_canvas = FigureCanvasQTAgg(self.crossplot_fig)
        self.crossplot_ax = self.crossplot_fig.add_subplot(111)
        self.crossplot_ax.set_facecolor('#2c2c2c')  # Set the axes background color
        self.crossplot_ax.tick_params(axis='x', colors='white')
        self.crossplot_ax.tick_params(axis='y', colors='white')
        self.crossplot_ax.spines['bottom'].set_color('white')
        self.crossplot_ax.spines['top'].set_color('white')
        self.crossplot_ax.spines['left'].set_color('white')
        self.crossplot_ax.spines['right'].set_color('white')
    
        # Connect heatmap click event
        self.heatmap_canvas.mpl_connect('button_press_event', self.on_heatmap_click)
    
        return left_panel


    def create_right_panel(self):
        """Right Panel with Correlation Threshold, Attribute Selection, Raw Data Table, and Export Button"""
        right_panel = QWidget()
        right_panel.setStyleSheet("""
            QWidget {
                background-color: #2c2c2c;
                color: #e0e0e0;
            }
            QLabel {
                color: #e0e0e0;
            }
            QLineEdit {
                background-color: #3c3f41;
                color: #e0e0e0;
                border: 1px solid #5a5a5a;
                border-radius: 4px;
                padding: 4px;
            }
            QLineEdit:focus {
                border: 1px solid #4a6984;
            }
        """)

        layout = QVBoxLayout(right_panel)

        # Correlation Threshold Input and Apply Button
        threshold_controls = QHBoxLayout()
        threshold_label = QLabel("Correlation Threshold:")
        self.threshold_input = QLineEdit()
        self.threshold_input.setPlaceholderText(str(self.threshold))
        self.threshold_input.setText(str(self.threshold))
        self.threshold_input.setFixedWidth(50)  # Small width
        apply_button = QPushButton("Apply")
        apply_button.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                color: white;
                padding: 6px 12px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
        """)
        apply_button.clicked.connect(self.apply_correlation_threshold)
        threshold_controls.addWidget(threshold_label)
        threshold_controls.addWidget(self.threshold_input)
        threshold_controls.addWidget(apply_button)
        layout.addLayout(threshold_controls)

        # Attribute Selector using DarkTwoListSelector
        self.attr_selector = DarkTwoListSelector(
            left_title="Important Attributes", 
            right_title="Selected Attributes"
        )
        layout.addWidget(self.attr_selector)

        # Raw Data Table (Bottom)
        layout.addWidget(QLabel("Raw Data Matrix"))
        self.raw_data_table = QTableWidget()
        self.raw_data_table.setStyleSheet("""
            QTableWidget {
                background-color: #3c3f41;
                color: #e0e0e0;
                border: 1px solid #5a5a5a;
            }
            QHeaderView::section {
                background-color: #2c2c2c;
                color: #e0e0e0;
                padding: 4px;
                border: 1px solid #5a5a5a;
                font-weight: bold;
            }
        """)
        self.populate_raw_data_table()
        layout.addWidget(self.raw_data_table)

        # Create horizontal layout for buttons
        button_layout = QHBoxLayout()

        # Export Button
        export_button = QPushButton("Export to Excel")
        export_button.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                padding: 6px 12px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        export_button.clicked.connect(self.export_to_excel)

        # Save to Regression Button
        save_regression_button = QPushButton("Save to Regression")
        save_regression_button.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                color: white;
                padding: 6px 12px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
        """)
        save_regression_button.clicked.connect(self.save_to_regression)

        # Add buttons to horizontal layout
        button_layout.addStretch()  # This pushes buttons to the right
        button_layout.addWidget(export_button)
        button_layout.addWidget(save_regression_button)

        # Add button layout to main layout
        layout.addLayout(button_layout)

        return right_panel





    def populate_attributes(self):
        """Ensure attributes from previous dialog are used"""
        print("Selected Attributes (raw):", self.selected_attrs)
    
        # Clear any existing items
        self.attr_selector.set_left_items([])
        self.attr_selector.set_right_items([])
    
        # Try different ways of populating attributes
        if isinstance(self.selected_attrs, list):
            # If it's already a list, use it directly
            self.attr_selector.set_left_items(self.selected_attrs)
        elif isinstance(self.selected_attrs, str):
            # If it's a single string, convert to list
            self.attr_selector.set_left_items([self.selected_attrs])
        else:
            # Try converting to list or use an empty list
            try:
                self.attr_selector.set_left_items(list(self.selected_attrs))
            except:
                print("Could not convert selected_attrs to list")
    
        # Verify items are added
        print("Left List Items:", self.attr_selector.get_left_items())

    def move_attributes_to_selected(self, x_col, y_col):
        """Moves both attributes from the important attributes list to the selected attributes list."""
        for attr in [x_col, y_col]:  # Loop through both attributes
            # Check if the attribute is already in the selected list
            already_selected = False
            for i in range(self.selected_attrs_list.count()):
                if self.selected_attrs_list.item(i).text() == attr:
                    already_selected = True
                    break

            if not already_selected:  # Add only if it's not already there
                self.selected_attrs_list.addItem(attr)

            # Find and remove from the "important attributes" list
            for i in range(self.important_attrs.count()):
                if self.important_attrs.item(i).text() == attr:
                    self.important_attrs.takeItem(i)  # Remove from important list
                    break  # Stop after removing to prevent index shifting issues


    def update_threshold(self):
        """Updates correlation threshold and filters attributes accordingly"""
        threshold = self.threshold_slider.value() / 100
        self.threshold_value.setText(f"{threshold:.2f}")

        # Remove attributes that fall below the threshold
        for i in range(self.important_attrs.count()):
            attr = self.important_attrs.item(i).text()
            attr_name = attr.split(".")[1]  # Extract actual column name

            if attr_name in self.correlation_matrix.columns:
                max_corr = self.correlation_matrix[attr_name].abs().max()
                if max_corr < threshold:
                    self.important_attrs.item(i).setHidden(True)
                else:
                    self.important_attrs.item(i).setHidden(False)

    def apply_correlation_threshold(self):
        """Moves attributes above the threshold from the important list to the selected list and updates the network graph."""
        try:
            # Get threshold from user input
            threshold = float(self.threshold_input.text().strip())
        except ValueError:
            QMessageBox.warning(self, "Invalid Input", "Please enter a valid numeric value for the threshold.")
            return

        # Set to track which attributes should be moved (ensure order is preserved)
        moved_attrs = set()

        # Iterate over the correlation matrix
        for i in range(len(self.correlation_matrix.index)):
            for j in range(i + 1, len(self.correlation_matrix.columns)):  # Only check upper triangle to avoid duplicates
                corr_value = abs(self.correlation_matrix.iloc[i, j])  # Get absolute correlation value

                if corr_value >= threshold:
                    x_col = self.correlation_matrix.columns[j]
                    y_col = self.correlation_matrix.index[i]
                    moved_attrs.add((x_col, y_col))  #   Store as a tuple (x_col, y_col)

        # Move identified attributes to the selected list
        for x_col, y_col in moved_attrs:  #   Ensure correct parameter order
            self.move_attributes_to_selected(x_col, y_col)  #   Now correctly sending both






    def move_multiple_attributes_to_selected(self, attributes):
        """Moves multiple attributes from the important list to the selected list at once."""
        for attr in attributes:
            # Ensure the attribute isn't already in the selected list
            already_selected = any(self.selected_attrs_list.item(i).text() == attr for i in range(self.selected_attrs_list.count()))
            if not already_selected:
                self.selected_attrs_list.addItem(attr)

            # Find and remove from the "important attributes" list
            for i in range(self.important_attrs.count()):
                if self.important_attrs.item(i).text() == attr:
                    self.important_attrs.takeItem(i)  # Remove from important list
                    break  # Stop after removing to prevent index shifting issues






    def draw_heatmap(self):
        """Draws an optimized heatmap with a smooth Red → Background Color → Green transition."""
        plt.close('all')  # Ensure clean slate
        self.heatmap_ax.clear()

        # **Set Dark Mode Background**
        bg_color = "#2C2C2C"  # Dark gray background (same as the UI theme)
        self.heatmap_fig.patch.set_facecolor(bg_color)
        self.heatmap_ax.set_facecolor(bg_color)

        # **Custom Colormap (E74C3C → BG Color → 2ECC71)**
        red_gray_green_cmap = mcolors.LinearSegmentedColormap.from_list(
            "red_gray_green_cmap", 
            ["#E74C3C", "#602020", bg_color, "#206040", "#2ECC71"], 
            N=256
        )

        # **Generate Heatmap with Custom Colors**
        heatmap = sns.heatmap(
            self.correlation_matrix, 
            annot=True, 
            cmap=red_gray_green_cmap,  # 🎨 Red → Background Gray → Green gradient
            center=0,
            fmt='.2f', 
            vmin=-1, vmax=1,  
            ax=self.heatmap_ax, 
            cbar=True, 
            square=False,
            robust=True,
            cbar_kws={'label': 'Correlation Coefficient', 'extend': 'both'},
            annot_kws={"size": 6, "fontweight": "bold"},
            xticklabels=True,
            yticklabels=True
        )

        plt.setp(self.heatmap_ax.get_xticklabels(), 
                 rotation=15, ha='center', va='top', rotation_mode='anchor',
                 fontsize=7, color='#e0e0e0')

        # **Fix Y-Axis Labels (-45° Rotation)**
        plt.setp(self.heatmap_ax.get_yticklabels(), 
                 rotation=-45, ha='right', va='center', rotation_mode='anchor',
                 fontsize=7, color='#e0e0e0')

        # **Customize Colorbar for Better Visibility**
        cbar = heatmap.collections[0].colorbar
        cbar.ax.set_ylabel('Correlation Coefficient', rotation=270, labelpad=15, color='#e0e0e0')

        # **Increase Colorbar Font Size**
        cbar.ax.tick_params(labelsize=9, colors='#e0e0e0', width=1.5, length=4)
        cbar.ax.yaxis.label.set_color('#e0e0e0')

        # **Customize Title and Labels**
        self.heatmap_ax.set_title("Correlation Heatmap", fontsize=10, color='#e0e0e0', pad=12)

        # **Adjust Axis Colors and Tick Parameters**
        self.heatmap_ax.tick_params(axis='both', colors='#e0e0e0')
        for spine in self.heatmap_ax.spines.values():
            spine.set_edgecolor('#e0e0e0')

        # **Highlight Strong Correlations with White/Black Text**
        for i in range(self.correlation_matrix.shape[0]):
            for j in range(self.correlation_matrix.shape[1]):
                value = self.correlation_matrix.iloc[i, j]
                color = 'white' if value < 0 else 'black'  # White for negative, black for positive
                if abs(value) > 0.5:  # Only highlight strong correlations
                    self.heatmap_ax.text(j + 0.5, i + 0.5, f'{value:.2f}', 
                                         ha='center', va='center', 
                                         color=color, fontsize=6, fontweight='bold')

        # **Ensure Proper Spacing & Draw**
        self.heatmap_fig.tight_layout()
        plt.subplots_adjust(bottom=0.3, left=0.25, right=0.95)  # Adjust space for long labels
        self.heatmap_canvas.draw()







    def draw_network(self):
        """Draw network graph of correlations in dark mode"""
        self.network_ax.clear()
        self.network_fig.patch.set_facecolor('#2c2c2c')
        self.network_ax.set_facecolor('#2c2c2c')
    
        try:
            # Get threshold from input
            threshold = self.threshold  # Default to 0.7 if empty
        
            # Create network graph
            G = nx.Graph()
        
            # Add edges for correlations above threshold
            for i in range(len(self.correlation_matrix.columns)):
                for j in range(i + 1, len(self.correlation_matrix.columns)):
                    corr = abs(self.correlation_matrix.iloc[i, j])
                    if corr >= threshold:
                        G.add_edge(
                            self.correlation_matrix.columns[i],
                            self.correlation_matrix.columns[j],
                            weight=corr
                        )
        
            if G.number_of_edges() == 0:
                self.network_ax.text(0.5, 0.5, "No correlations above threshold",
                                   ha='center', va='center', color='#e0e0e0', fontsize=8)
                self.network_canvas.draw()
                return
        
            # Calculate layout
            pos = nx.spring_layout(G, k=1, iterations=50)
        
            # Draw nodes
            nx.draw_networkx_nodes(G, pos,
                                 node_color='#4a6984',
                                 node_size=500,  # Reduced node size
                                 ax=self.network_ax)
        
            # Draw edges with varying thickness based on correlation
            edge_weights = [G[u][v]['weight'] * 1.5 for u, v in G.edges()]
            nx.draw_networkx_edges(G, pos,
                                 width=edge_weights,
                                 edge_color='#5a5a5a',
                                 ax=self.network_ax)
        
            # Add labels
            nx.draw_networkx_labels(G, pos,
                                  font_size=6,  # Smaller font
                                  ax=self.network_ax,
                                  font_color='#e0e0e0')
        
            # Add edge labels (correlation values)
            edge_labels = {(u, v): f'{G[u][v]["weight"]:.2f}'
                         for u, v in G.edges()}
            nx.draw_networkx_edge_labels(G, pos,
                                       edge_labels=edge_labels,
                                       font_size=5,  # Even smaller font
                                       font_color='#e0e0e0')
        
            self.network_ax.set_title(f"Correlation Network", 
                                      color='#e0e0e0', fontsize=10)
        
        except ValueError as e:
            self.network_ax.text(0.5, 0.5, "Error: Invalid threshold value",
                               ha='center', va='center', color='#e0e0e0', fontsize=8)
    
        # Remove axis
        self.network_ax.set_axis_off()
    
        # Update canvas
        self.network_canvas.draw()


    def show_context_menu(self, position):
        menu = QMenu()
        send_action = menu.addAction("Send to Selected")
        action = menu.exec_(self.important_attrs.mapToGlobal(position))
        if action == send_action:
            self.add_attribute()

    def filter_attributes(self, text):
        for i in range(self.important_attrs.count()):
            item = self.important_attrs.item(i)
            item.setHidden(text.lower() not in item.text().lower())

    def add_attribute(self):
        """Move selected items from important_attrs to selected_attrs_list"""
        items = self.important_attrs.selectedItems()
        for item in items:
            # Take the item from important_attrs
            self.important_attrs.takeItem(self.important_attrs.row(item))
            # Add to selected_attrs_list (not selected_attrs)
            self.selected_attrs_list.addItem(item.text())

    def remove_attribute(self):
        """Move selected items from selected_attrs_list back to important_attrs"""
        items = self.selected_attrs_list.selectedItems()
        for item in items:
            # Take the item from selected_attrs_list
            self.selected_attrs_list.takeItem(self.selected_attrs_list.row(item))
            # Add back to important_attrs
            self.important_attrs.addItem(item.text())





    def on_heatmap_click(self, event):
        """Handles left-click for crossplot and right-click to move attributes to the selected list."""
        if event.inaxes == self.heatmap_ax:  # Ensure click is inside the heatmap
            try:
                col = round(event.xdata - 0.5)  # Get column index
                row = round(event.ydata - 0.5)  # Get row index

                # Validate indices and avoid diagonal self-correlation
                if (0 <= row < len(self.correlation_matrix.index) and 
                    0 <= col < len(self.correlation_matrix.columns) and 
                    row != col):

                    x_col = self.correlation_matrix.columns[col]  # Get X-axis attribute
                    y_col = self.correlation_matrix.index[row]    # Get Y-axis attribute

                    if event.button == 1:  # Left-click → Update Crossplot
                        self.update_crossplot(row, col)

                    elif event.button == 3:  # Right-click → Move attributes to selection
                        self.move_attributes_to_selected(x_col, y_col)

            except (TypeError, ValueError, IndexError) as e:
                print(f"Heatmap click error: {e}")


    def switch_view(self, view_type):
        """Switch between correlation matrix and network graph views"""
        if view_type == "Correlation Matrix":
            self.viz_stack.setCurrentIndex(0)
        elif view_type == "Network Graph":
            self.viz_stack.setCurrentIndex(1)

    def update_crossplot(self, row, col):
        """Updates the crossplot with the selected variables."""
        x_col = self.correlation_matrix.columns[col]
        y_col = self.correlation_matrix.index[row]

        self.crossplot_ax.clear()
        self.crossplot_fig.patch.set_facecolor('#2c2c2c')
        self.crossplot_ax.set_facecolor('#2c2c2c')

        self.crossplot_ax.set_title(f"Crossplot: {x_col} vs {y_col}", 
                                    color='#e0e0e0', fontsize=10)
        self.crossplot_ax.set_xlabel(x_col, color='#e0e0e0', fontsize=8)
        self.crossplot_ax.set_ylabel(y_col, color='#e0e0e0', fontsize=8)

        # Set tick colors
        self.crossplot_ax.tick_params(colors='#e0e0e0', labelsize=8)
        for spine in self.crossplot_ax.spines.values():
            spine.set_edgecolor('#5a5a5a')

        sns.regplot(
            x=self.data_matrix[x_col], 
            y=self.data_matrix[y_col], 
            scatter_kws={'alpha': 0.7, 'color': '#00aaff', 's': 30},  # Bright blue scatter points
            line_kws={'color': '#e74c3c', 'linewidth': 1},  # Thinner line
            ax=self.crossplot_ax
        )

        # Display correlation coefficient
        corr = self.data_matrix[x_col].corr(self.data_matrix[y_col])
        self.crossplot_ax.annotate(f'Correlation: {corr:.2f}', 
                                   xy=(0.05, 0.95), 
                                   xycoords='axes fraction',
                                   fontsize=8,
                                   color='#e0e0e0',
                                   verticalalignment='top')

        self.crossplot_canvas.draw()


    def populate_raw_data_table(self):
        """Fills the raw data table and resizes columns to fit numeric values, not headers."""
        self.raw_data_table.setRowCount(len(self.data_matrix.index))
        self.raw_data_table.setColumnCount(len(self.data_matrix.columns))
        self.raw_data_table.setHorizontalHeaderLabels(self.data_matrix.columns)
        self.raw_data_table.setVerticalHeaderLabels(self.data_matrix.index)

        for i in range(len(self.data_matrix.index)):
            for j in range(len(self.data_matrix.columns)):
                value = self.data_matrix.iloc[i, j]
                item = QTableWidgetItem(f"{value:.3f}" if isinstance(value, float) else str(value))
                item.setTextAlignment(Qt.AlignCenter)
                self.raw_data_table.setItem(i, j, item)

        # Set column width to a smaller, fixed size
        for col in range(self.raw_data_table.columnCount()):
            self.raw_data_table.setColumnWidth(col, 80)  # Adjust this value as needed
        
        # Ensure headers are visible
        self.raw_data_table.horizontalHeader().setStretchLastSection(False)


    def export_to_excel(self):
        """Export correlation matrix to an Excel file with colors"""
        try:
            file_path, _ = QFileDialog.getSaveFileName(self, "Save File", "", "Excel Files (*.xlsx)")
            if not file_path:
                return  # User canceled

            if not file_path.lower().endswith('.xlsx'):
                file_path += '.xlsx'

            workbook = openpyxl.Workbook()
            std_sheet = workbook.active
            workbook.remove(std_sheet)

            sheet = workbook.create_sheet("Correlation Matrix")

            # Write headers
            for col_idx, col_name in enumerate(self.correlation_matrix.columns, start=2):
                sheet.cell(row=1, column=col_idx, value=col_name)

            for row_idx, row_name in enumerate(self.correlation_matrix.index, start=2):
                sheet.cell(row=row_idx, column=1, value=row_name)

                for col_idx, col_name in enumerate(self.correlation_matrix.columns, start=2):
                    value = self.correlation_matrix.loc[row_name, col_name]
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)

                    # Apply color formatting based on correlation value
                    if pd.notna(value):
                        qcolor = self.get_seaborn_qcolor(value)
                        rgb = qcolor.getRgb()[:3]
                        hex_color = f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"
                        cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

            workbook.save(file_path)
            QMessageBox.information(self, "Export Successful", f"File saved to {file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"An error occurred while exporting: {str(e)}")

    def get_seaborn_qcolor(self, value):
        """Convert Seaborn heatmap color to QColor for Qt tables"""
        if pd.isna(value):
            return QColor(255, 255, 255)  # White for NaN
        cmap = sns.color_palette("RdYlBu", as_cmap=True)
        norm = mcolors.Normalize(vmin=-1, vmax=1)
        rgba = cmap(norm(value))
        rgb = [int(255 * c) for c in rgba[:3]]
        return QColor(*rgb)

    def save_to_regression(self):
        """Save correlation results to a regression table"""
        try:
            # Show dialog to get table name
            save_dialog = SaveRegressionDialog(self.parent().db_manager, self)
            if save_dialog.exec() != QDialog.Accepted:
                return
            
            table_name = save_dialog.get_selected_table()
            description = save_dialog.get_description()
            replace_mode = save_dialog.get_save_mode()
            
            if not table_name:
                QMessageBox.warning(self, "Invalid Input", "Please select or enter a table name")
                return
            
            # Get selected attributes from the selected_attrs_list
            selected_attributes = self.attr_selector.get_right_items()
            
            if len(selected_attributes) == 0:
                QMessageBox.warning(self, "No Attributes", "Please select attributes to save")
                return
            
            # Filter data matrix to only include selected attributes
            filtered_data = self.data_matrix[selected_attributes].copy()

            # Ensure UWI is a column
            if filtered_data.index.name == 'UWI':
                filtered_data = filtered_data.reset_index()

            # Remove duplicates by keeping the last entry for each UWI
            filtered_data = filtered_data.drop_duplicates(subset=['UWI'], keep='last')

            print(filtered_data)
            # Use db_manager to save data
            self.parent().db_manager.save_correlation_to_regression(
                table_name, 
                filtered_data,
                description,
                replace_mode
            )
            
            QMessageBox.information(self, "Success", f"Data {'replaced in' if replace_mode else 'added to'} regression table: {table_name}")
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error saving to regression: {str(e)}")



class GenerateCorrelationMatrix(QDialog):
    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.setWindowTitle("Well Zone Correlation Analysis")
        self.setMinimumWidth(800)
        self.setMinimumHeight(600)
        
        # Overall layout
        layout = QVBoxLayout()
        
        # Create a two-list selector for UWIs using your TwoListSelector
        self.uwi_selector = TwoListSelector(
            left_title="Available UWIs", 
            right_title="Selected UWIs"
        )
        
        # Create a two-list selector for Attributes
        self.attr_selector = TwoListSelector(
            left_title="Available Attributes", 
            right_title="Selected Attributes"
        )
        
        # Add selectors to layout
        splitter = QSplitter(Qt.Vertical)
        splitter.addWidget(self.uwi_selector)
        splitter.addWidget(self.attr_selector)
        layout.addWidget(splitter)
        
        bottom_layout = QHBoxLayout()

        # Left side: Ignore zeros and threshold
        left_controls_layout = QHBoxLayout()

        # Ignore zeros checkbox
        self.ignore_zeros_checkbox = QCheckBox("Ignore Zeros")
        left_controls_layout.addWidget(self.ignore_zeros_checkbox)

        # Threshold selection
        threshold_label = QLabel("Threshold:")
        threshold_label.setFixedWidth(70)
        left_controls_layout.addWidget(threshold_label)

        self.threshold_input = QComboBox()
        self.threshold_input.setFixedWidth(70)
        threshold_values = [f"{i/20:.2f}" for i in range(21)]
        self.threshold_input.addItems(threshold_values)
        self.threshold_input.setCurrentIndex(0)
        left_controls_layout.addWidget(self.threshold_input)

        # Add the left controls to the bottom layout
        bottom_layout.addLayout(left_controls_layout)

        # Add stretch to push run button to the right
        bottom_layout.addStretch()

        # Run button with green styling
        btn_run = QPushButton("Run")
        btn_run.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                color: white;
                padding: 6px 12px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
        """)
        btn_run.clicked.connect(self.run_analysis)
        bottom_layout.addWidget(btn_run)

        # Add the bottom layout to the main layout
        layout.addLayout(bottom_layout)

        self.setLayout(layout)
        self.load_data()


        
    def _is_numeric_column(self, data, col_idx):
        try:
            print(f"\nChecking if column is numeric:")
            for row_idx, row in enumerate(data):
                value = row[col_idx]
            
                # Skip empty values
                if value is None or value == "":
                    continue

                # If it's already a float or int, it's numeric
                if isinstance(value, (float, int)):
                    return True

                # Try to convert any strings
                if isinstance(value, str):
                    value = value.strip()
                    try:
                        float(value)
                        return True
                    except ValueError:
                        return False

            return False
        except Exception as e:
            print(f"Error checking numeric: {e}")
            return False


    def load_data(self):
        """Load available UWIs and attributes from the database"""
        try:
            # Define columns to exclude
            columns_to_exclude = [
                'Zone_Name', 'Zone_Type', 'Attribute_Type',
                'Top_Depth', 'Base_Depth', 'UWI',
                'Top_X_Offset', 'Base_X_Offset', 'Top_Y_Offset', 'Base_Y_Offset',
                'Angle_Top', 'Angle_Base', 'Total_Lateral_Length'
            ]
    
            # Load UWIs 
            print("\nGetting UWIs...")
            UWIs = self.db_manager.get_UWIs()
            print(f"Found UWIs: {UWIs}")
            self.uwi_selector.set_left_items(UWIs)

            # Fetch well zones
            print("\nFetching Well Zones...")
            well_zones = self.db_manager.fetch_zone_names_by_type("Well")
            print(f"Found Well Zones: {[zone[0] for zone in well_zones]}")

            numeric_columns = []

            # Process each well zone
            for zone_tuple in well_zones:
                zone_name = zone_tuple[0]
                print(f"\nProcessing Zone: {zone_name}")

                try:
                    # Get the table name using the database manager method
                    zone_table_name = self.db_manager.get_table_name_from_zone(zone_name)
                    print(f"Using table: {zone_table_name}")

                    # Fetch zone table data
                    data, columns = self.db_manager.fetch_zone_table_data(zone_table_name)
                    print(f"Columns in {zone_table_name}: {columns}")

                    # Find numeric columns
                    for col in columns:  # Process all columns
                        # Skip excluded columns and UWI
                        if col.upper() == 'UWI' or any(exclude.lower() in col.lower() for exclude in columns_to_exclude):
                            print(f"Skipping excluded column: {col}")
                            continue
                    
                        col_idx = columns.index(col)
                        print(f"Checking Column: {col}")

                        if self._is_numeric_column(data, col_idx):
                            # Add the column with the full table name
                            numeric_columns.append(f"{zone_table_name}.{col}")
                            print(f"Added numeric column: {zone_table_name}.{col}")
                        else:
                            print(f"Column {col} rejected - not numeric")

                except Exception as e:
                    print(f"Error processing zone {zone_name}: {str(e)}")
                    continue

            print(f"\nFinal List of Numeric Attributes: {numeric_columns}")
            self.attr_selector.set_left_items(numeric_columns)

        except Exception as e:
            print(f"Error loading data: {str(e)}")
            QMessageBox.critical(self, "Error", f"Error loading data: {str(e)}")
            


    def validate_numeric_data(self, data):
        """Ensure all data is in correct format for correlation"""
        valid_data = pd.DataFrame()
        
        for column in data.columns:
            # Convert to numeric, forcing errors to NaN
            series = pd.to_numeric(data[column], errors='coerce')
            
            # Only include columns with actual numeric data
            if not series.isna().all():
                valid_data[column] = series
                
            # Print debugging info
            print(f"\nValidating column: {column}")
            print(f"Original dtype: {data[column].dtype}")
            print(f"Converted dtype: {series.dtype}")
            print(f"Number of valid values: {series.count()}")
        
        return valid_data




    def debug_data(self, numeric_data):
        """Debug data issues before correlation"""
        print("\nData Debug Information:")
        for col in numeric_data.columns:
            print(f"\nColumn: {col}")
            print(f"Type: {type(numeric_data[col])}")
            print(f"Shape: {numeric_data[col].shape}")
            print(f"Sample: {numeric_data[col].head()}")
            
            # Check if column is valid for correlation
            if not isinstance(numeric_data[col], (pd.Series, np.ndarray)):
                print(f"WARNING: Column {col} is not in correct format")
                # Convert to proper format
                numeric_data[col] = pd.Series(numeric_data[col])
        
        return numeric_data

            
    def run_analysis(self):
        """Run correlation analysis on selected items"""
        try:
            # Get selected UWIs and attributes
            UWIs = self.uwi_selector.get_right_items()
            selected_attrs = self.attr_selector.get_right_items()

            print("Selected UWIs:", UWIs)
            print("Selected Attributes:", selected_attrs)

            if not UWIs or not selected_attrs:
                QMessageBox.warning(self, "Invalid Selection", "Please select at least one UWI and attribute")
                return

            #   Get threshold value
            try:
                threshold = float(self.threshold_input.currentText().strip())
            except ValueError:
                QMessageBox.warning(self, "Invalid Input", "Please enter a valid numeric value for the threshold.")
                return

            #   Call `fetch_correlation_data` from `db_manager`
            results = self.db_manager.fetch_correlation_data(UWIs, selected_attrs)

            if not results:
                QMessageBox.warning(self, "No Data", "No data found for the selected combinations")
                return

            # Create DataFrame with all data
            columns = ['UWI'] + [attr for attr in selected_attrs if attr.split('.')[1].upper() != 'UWI']
            df = pd.DataFrame(results, columns=columns)
            df = df.set_index('UWI')

            # Convert to numeric, handling errors
            numeric_df = df.apply(pd.to_numeric, errors='coerce')

            # Handle zero values if checkbox is checked
            if self.ignore_zeros_checkbox.isChecked():
                print("Ignoring zero values in correlation calculation")
                numeric_df = numeric_df.replace(0, np.nan)

            print("\nFinal numeric data:")
            print(f"Shape: {numeric_df.shape}")
            print(f"Columns: {numeric_df.columns.tolist()}")

            #   Calculate correlation matrix
            corr_matrix = numeric_df.corr(method='pearson')

            #   Apply threshold filter
            filtered_corr_matrix = corr_matrix.copy()
            filtered_corr_matrix[abs(filtered_corr_matrix) < threshold] = None  # Hide weak correlations

            print("\nFiltered correlation matrix:")
            print(filtered_corr_matrix)

            #   Pass the threshold to CorrelationDisplayDialog
            dialog = CorrelationDisplayDialog(filtered_corr_matrix, numeric_df, selected_attrs, threshold, self)
            dialog.exec()

        except Exception as e:
            import traceback
            print("Error details:", str(e))
            print("Full traceback:", traceback.format_exc())
            QMessageBox.critical(self, "Error", f"Error running analysis: {str(e)}")


# Example usage and testing:
if __name__ == '__main__':
    import sys
    from PySide6.QtWidgets import QApplication
    
    class MockDBManager:
        def connect(self):
            pass
            
        def disconnect(self):
            pass
            
        def get_UWIs(self):
            return ['UWI1', 'UWI2', 'UWI3', 'UWI4', 'UWI5']
            
        def fetch_zone_names_by_type(self, zone_type):
            return [('Zone1',), ('Zone2',)]
            
        def fetch_zone_table_data(self, zone_name):
            # Mock data
            data = [
                ['UWI1', 100.0, 200.0, 300.0],
                ['UWI2', 150.0, 250.0, 350.0],
                ['UWI3', 175.0, 225.0, 375.0],
                ['UWI4', 125.0, 275.0, 325.0],
                ['UWI5', 160.0, 240.0, 360.0]
            ]
            columns = ['UWI', 'Value1', 'Value2', 'Value3']
            return data, columns

    app = QApplication(sys.argv)
    db_manager = MockDBManager()
    dialog = GenerateCorrelationMatrix(db_manager)
    dialog.show()
    sys.exit(app.exec())